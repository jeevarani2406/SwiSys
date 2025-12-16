import io
import logging
import traceback
import os
import re
import json
from collections import defaultdict
from datetime import datetime, date
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.core.files.storage import default_storage
from django.http import JsonResponse

from openpyxl import load_workbook

from .models import Vehicle, SPN, PGN, VehicleSPN, VehiclePGN, StandardFile, AuxiliaryFile, Category, J1939ParameterDefinition
from rest_framework import generics
from .serializers import (
    VehicleSerializer, VehicleSPNSerializer, StandardFileSerializer, AuxiliaryFileSerializer, 
    CategorySerializer, PGNSerializer, SPNSerializer, J1939ParameterDefinitionSerializer,
    SPNDecodeRequestSerializer, SPNDecodeResponseSerializer
)
from django.db.models import Count
from rest_framework.parsers import MultiPartParser, FormParser

# Import pandas lazily inside methods to avoid import-time failures during
# Django management commands (makemigrations/migrate) when pandas may not be
# available in the environment. When available we will use it for Excel
# parsing; otherwise we fallback to openpyxl-only parsing.
pd = None
try:
    import pandas as pd  # type: ignore
except Exception:
    pd = None

logger = logging.getLogger(__name__)


@method_decorator(csrf_exempt, name='dispatch')
class J1939UploadView(APIView):
    """
    POST /api/j1939/upload
    
    Accepts multiple Excel files from the frontend, parses their content, and extracts
    structured vehicle, PGN, and SPN information.
    
    Returns:
    {
        "status": "success",
        "vehicles": [
            {
                "name": "Truck A",
                "brand": "Volvo",
                "pgns": [601, 602],
                "spns": [
                    { "pgn": 601, "spn": 190, "description": "Engine Speed" },
                    { "pgn": 602, "spn": 247, "description": "Oil Temp" }
                ]
            }
        ],
        "errors": []
    }
    """
    permission_classes = [permissions.AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, format=None):
        """
        Process multiple Excel file uploads.
        """
        # Gather files from common field names
        files = []
        if request.FILES.getlist('file'):
            files = request.FILES.getlist('file')
        elif request.FILES.getlist('files[]'):
            files = request.FILES.getlist('files[]')
        elif request.FILES.getlist('files'):
            files = request.FILES.getlist('files')
        else:
            try:
                files = list(request.FILES.values())
            except Exception:
                files = []

        if not files:
            return Response({
                'status': 'error',
                'vehicles': [],
                'errors': ['No files uploaded']
            }, status=status.HTTP_400_BAD_REQUEST)

        vehicles = []
        errors = []
        today = timezone.now().date()

        for f in files:
            fname = getattr(f, 'name', '<unknown>')
            logger.info('Processing file: %s', fname)
            
            # Initialize PGN count variables for this file
            total_pgn_count = 0
            unique_pgn_count = 0
            unique_pgn_list = []
            # All file types are now accepted - we'll detect the format automatically
            # Files without extensions or unknown extensions will be treated as CSV/text

            try:
                # Read file content
                f.seek(0)
                file_content = f.read()
                f.seek(0)

                # Parse file (Excel or text-based)
                try:
                    # Only .xlsx and .xls are treated as Excel files
                    # Everything else (including files without extensions) is treated as CSV/text
                    excel_extensions = ('.xlsx', '.xls')
                    is_excel_file = fname.lower().endswith(excel_extensions)
                    
                    if not is_excel_file:
                        # CSV/TXT/LOG: read as single-sheet structure with multi-encoding support
                        f.seek(0)
                        
                        # Try multiple encodings for CSV files
                        encodings_to_try = ['utf-8', 'utf-8-sig', 'gb2312', 'gbk', 'gb18030', 
                                           'big5', 'utf-16', 'utf-16-le', 'latin1', 'cp1252', 'iso-8859-1']
                        
                        df = None
                        decoded_text = None
                        encoding_used = None
                        
                        if pd is not None:
                            # Try pandas with multiple encodings
                            for encoding in encodings_to_try:
                                try:
                                    df = pd.read_csv(io.BytesIO(file_content), encoding=encoding)
                                    encoding_used = encoding
                                    logger.info(f"CSV {fname} parsed successfully with encoding: {encoding}")
                                    break
                                except (UnicodeDecodeError, UnicodeError):
                                    continue
                                except Exception as enc_err:
                                    # Try next encoding
                                    continue
                            
                            # Last resort: use latin1 which accepts any byte
                            if df is None:
                                try:
                                    df = pd.read_csv(io.BytesIO(file_content), encoding='latin1', on_bad_lines='skip')
                                    encoding_used = 'latin1-fallback'
                                    logger.info(f"CSV {fname} parsed with latin1 fallback")
                                except Exception as final_err:
                                    logger.error(f"All encoding attempts failed for {fname}: {final_err}")
                                    raise
                            
                            # Calculate PGN counts using pandas (matching the exact Python logic)
                            # Filter for main message rows where Index is not NaN
                            total_pgn_count = 0
                            unique_pgn_count = 0
                            unique_pgn_list = []
                            
                            if df is not None and 'Index' in df.columns and 'PGN(H)' in df.columns:
                                # Filter for rows where Index is not NaN (main message rows only)
                                message_df = df[df['Index'].notna()]
                                # Total: count of non-null PGN(H) values in filtered rows
                                total_pgn_count = int(message_df['PGN(H)'].count())
                                # Unique: count of distinct PGN(H) values
                                unique_pgn_count = int(message_df['PGN(H)'].nunique())
                                # Get list of unique PGN values
                                unique_pgn_list = message_df['PGN(H)'].dropna().unique().tolist()
                                unique_pgn_list = [str(x).upper() for x in unique_pgn_list if pd.notna(x)]
                                logger.info(f"PGN counts for {fname}: Total={total_pgn_count}, Unique={unique_pgn_count}")
                            
                            df_dict = {os.path.splitext(fname)[0]: df}
                        else:
                            # Fallback CSV parser with multi-encoding support
                            for encoding in encodings_to_try:
                                try:
                                    decoded_text = file_content.decode(encoding)
                                    encoding_used = encoding
                                    break
                                except (UnicodeDecodeError, UnicodeError):
                                    continue
                            
                            # Final fallback
                            if decoded_text is None:
                                decoded_text = file_content.decode('latin1', errors='replace')
                                encoding_used = 'latin1-fallback'
                            
                            import csv as _csv
                            rows = list(_csv.reader(decoded_text.splitlines()))
                            if not rows:
                                df_dict = {os.path.splitext(fname)[0]: {}}
                            else:
                                headers = rows[0]
                                data_rows = rows[1:]
                                sheet_data = {}
                                for col_idx, header in enumerate(headers):
                                    sheet_data[header] = [r[col_idx] if col_idx < len(r) else None for r in data_rows]
                                df_dict = {os.path.splitext(fname)[0]: sheet_data}
                    else:
                        # Excel file (.xlsx, .xls)
                        if pd is not None:
                            # Use pandas for better Excel parsing
                            # Try openpyxl first for .xlsx, xlrd for .xls
                            try:
                                if fname.lower().endswith('.xls') and not fname.lower().endswith('.xlsx'):
                                    # Old .xls format - try xlrd engine
                                    try:
                                        df_dict = pd.read_excel(io.BytesIO(file_content), sheet_name=None, engine='xlrd')
                                    except Exception:
                                        # Fallback to openpyxl
                                        df_dict = pd.read_excel(io.BytesIO(file_content), sheet_name=None, engine='openpyxl')
                                else:
                                    # .xlsx format
                                    df_dict = pd.read_excel(io.BytesIO(file_content), sheet_name=None, engine='openpyxl')
                                
                                logger.info(f"Excel {fname} parsed successfully")
                                
                                # Calculate PGN counts for Excel files (same logic as CSV)
                                for sheet_name, sheet_df in df_dict.items():
                                    if sheet_df is not None and not sheet_df.empty:
                                        if 'Index' in sheet_df.columns and 'PGN(H)' in sheet_df.columns:
                                            # Filter for rows where Index is not NaN (main message rows only)
                                            message_df = sheet_df[sheet_df['Index'].notna()]
                                            # Total: count of non-null PGN(H) values in filtered rows
                                            total_pgn_count = int(message_df['PGN(H)'].count())
                                            # Unique: count of distinct PGN(H) values
                                            unique_pgn_count = int(message_df['PGN(H)'].nunique())
                                            # Get list of unique PGN values
                                            unique_pgn_list = message_df['PGN(H)'].dropna().unique().tolist()
                                            unique_pgn_list = [str(x).upper() for x in unique_pgn_list if pd.notna(x)]
                                            logger.info(f"PGN counts for {fname} (sheet: {sheet_name}): Total={total_pgn_count}, Unique={unique_pgn_count}")
                                            break  # Use first sheet with valid data
                                            
                            except Exception as excel_err:
                                logger.error(f"Excel parsing error for {fname}: {excel_err}")
                                raise
                        else:
                            # Fallback to openpyxl - convert to simple data structure
                            wb = load_workbook(filename=io.BytesIO(file_content), data_only=True)
                            df_dict = {}
                            for sheet_name in wb.sheetnames:
                                sheet = wb[sheet_name]
                                # Convert sheet to list of lists for processing
                                data = []
                                headers = None
                                for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                                    if row_idx == 0:
                                        headers = [str(cell) if cell is not None else f'Col{i}' for i, cell in enumerate(row)]
                                    else:
                                        data.append(list(row))
                                # Create a simple dict-like structure for compatibility
                                if headers and data:
                                    # Store as dict with column names as keys
                                    sheet_data = {}
                                    for col_idx, header in enumerate(headers):
                                        sheet_data[header] = [row[col_idx] if col_idx < len(row) else None for row in data]
                                    df_dict[sheet_name] = sheet_data
                                else:
                                    df_dict[sheet_name] = {}
                except Exception as parse_exc:
                    error_msg = f'Failed to parse file: {str(parse_exc)}'
                    errors.append({
                        'filename': fname,
                        'error': error_msg
                    })
                    logger.error('File parsing error for %s: %s', fname, str(parse_exc), exc_info=True)
                    continue

                # Extract vehicle information
                vehicle_name = None
                brand = None
                pgns = set()
                spns_data = {}  # {(pgn, spn): description}

                # Try to extract from all sheets
                for sheet_name, df in df_dict.items():
                    # Handle both pandas DataFrame and dict structure
                    if df is None:
                        continue
                    
                    # Check if it's a pandas DataFrame
                    is_dataframe = pd is not None and isinstance(df, pd.DataFrame)
                    if is_dataframe and df.empty:
                        continue
                    
                    # For dict structure, check if it's empty
                    if not is_dataframe and (not df or len(df) == 0):
                        continue

                    # Extract vehicle name and brand
                    vehicle_name_aliases = ['vehicle name', 'veh name', 'vehicle', 'veh', 'unit name', 'name']
                    brand_aliases = ['brand', 'make', 'manufacturer', 'manufacturer name']

                    # Search in first few rows and columns
                    max_rows = len(df) if is_dataframe else max([len(v) for v in df.values()] if isinstance(df, dict) else [0])
                    max_cols = len(df.columns) if is_dataframe else len(df) if isinstance(df, dict) else 0
                    
                    for row_idx in range(min(10, max_rows)):
                        for col_idx in range(min(10, max_cols)):
                            try:
                                if is_dataframe:
                                    cell_value = str(df.iloc[row_idx, col_idx]).strip().lower()
                                else:
                                    # For dict structure, access by column name
                                    col_names = list(df.keys()) if isinstance(df, dict) else []
                                    if col_idx < len(col_names):
                                        col_name = col_names[col_idx]
                                        cell_value = str(df[col_name][row_idx] if row_idx < len(df[col_name]) else '').strip().lower()
                                    else:
                                        continue
                                
                                # Check for vehicle name
                                if not vehicle_name:
                                    for alias in vehicle_name_aliases:
                                        if alias in cell_value:
                                            # Try to get value from adjacent cell or next row
                                            try:
                                                if col_idx + 1 < max_cols:
                                                    if is_dataframe:
                                                        candidate = str(df.iloc[row_idx, col_idx + 1]).strip()
                                                    else:
                                                        next_col = col_names[col_idx + 1] if col_idx + 1 < len(col_names) else None
                                                        candidate = str(df[next_col][row_idx] if next_col and row_idx < len(df[next_col]) else '').strip()
                                                    if candidate and candidate.lower() not in ['nan', 'none', '']:
                                                        vehicle_name = candidate
                                                        break
                                            except:
                                                pass
                                
                                # Check for brand
                                if not brand:
                                    for alias in brand_aliases:
                                        if alias in cell_value:
                                            try:
                                                if col_idx + 1 < max_cols:
                                                    if is_dataframe:
                                                        candidate = str(df.iloc[row_idx, col_idx + 1]).strip()
                                                    else:
                                                        next_col = col_names[col_idx + 1] if col_idx + 1 < len(col_names) else None
                                                        candidate = str(df[next_col][row_idx] if next_col and row_idx < len(df[next_col]) else '').strip()
                                                    if candidate and candidate.lower() not in ['nan', 'none', '']:
                                                        brand = candidate
                                                        break
                                            except:
                                                pass
                            except:
                                continue

                    # Extract PGNs and SPNs
                    # Priority: 'PGN(H)' column for hex PGN values, then other aliases
                    pgn_h_aliases = ['pgn(h)', 'pgn_h', 'pgn h', 'pgn(hex)', 'pgn_hex']
                    pgn_aliases = ['pgn', 'pgn number', 'pgn_no', 'pgn_number', 'parameter group number']
                    spn_aliases = ['spn', 'spn number', 'spn_no', 'spn_number', 'suspect parameter number']
                    index_aliases = ['index', 'idx', 'no', 'no.', 'row', 'row_no']

                    # Find column indices/names
                    index_col_idx = None  # For Index column to identify main message rows
                    index_col_name = None
                    pgn_h_col_idx = None  # For PGN(H) hex column
                    pgn_h_col_name = None
                    pgn_col_idx = None
                    pgn_col_name = None
                    spn_col_idx = None
                    spn_col_name = None
                    desc_col_idx = None
                    desc_col_name = None

                    # Lists to track PGN(H) values for counting
                    all_pgn_h_values = []  # All non-empty PGN(H) values (for total count)
                    unique_pgn_h_values = set()  # Unique PGN(H) values

                    # Check header row
                    if is_dataframe:
                        columns = df.columns
                        for col_idx, col_name in enumerate(columns):
                            col_str = str(col_name).strip().lower()
                            # Check for Index column (to identify main message rows)
                            if col_str in index_aliases or col_str == 'index':
                                index_col_idx = col_idx
                            # Check for PGN(H) column first (priority)
                            if any(alias == col_str for alias in pgn_h_aliases):
                                pgn_h_col_idx = col_idx
                            elif any(alias in col_str for alias in pgn_aliases):
                                pgn_col_idx = col_idx
                            if any(alias in col_str for alias in spn_aliases):
                                spn_col_idx = col_idx
                            if 'description' in col_str or 'desc' in col_str:
                                desc_col_idx = col_idx
                    else:
                        # For dict structure
                        for col_name in df.keys():
                            col_str = str(col_name).strip().lower()
                            # Check for Index column
                            if col_str in index_aliases or col_str == 'index':
                                index_col_name = col_name
                            # Check for PGN(H) column first (priority)
                            if any(alias == col_str for alias in pgn_h_aliases):
                                pgn_h_col_name = col_name
                            elif any(alias in col_str for alias in pgn_aliases):
                                pgn_col_name = col_name
                            if any(alias in col_str for alias in spn_aliases):
                                spn_col_name = col_name
                            if 'description' in col_str or 'desc' in col_str:
                                desc_col_name = col_name

                    # Extract data from rows
                    num_rows = len(df) if is_dataframe else max([len(v) for v in df.values()] if isinstance(df, dict) else [0])
                    current_pgn = None  # Track current PGN for rows without explicit PGN
                    
                    for row_idx in range(num_rows):
                        try:
                            pgn_value = None
                            pgn_h_value = None  # For PGN(H) hex value
                            spn_value = None
                            desc_value = ''
                            
                            # Check if this is a main message row (Index column has value)
                            # In J1939 log files, only rows with Index value are main messages
                            # Detail rows (SPNs) have NaN in the Index column
                            is_main_message_row = True  # Default to True if no Index column
                            
                            if is_dataframe and index_col_idx is not None:
                                try:
                                    index_val = df.iloc[row_idx, index_col_idx]
                                    is_main_message_row = pd is not None and pd.notna(index_val)
                                except:
                                    pass
                            elif not is_dataframe and index_col_name and index_col_name in df:
                                try:
                                    index_val = df[index_col_name][row_idx] if row_idx < len(df[index_col_name]) else None
                                    is_main_message_row = index_val is not None and str(index_val).strip().lower() not in ['', 'nan', 'none', 'null']
                                except:
                                    pass

                            # Get PGN(H) - hex PGN column (priority)
                            # Only count for PGN stats if this is a main message row
                            if is_dataframe:
                                if pgn_h_col_idx is not None:
                                    try:
                                        pgn_h_val = df.iloc[row_idx, pgn_h_col_idx]
                                        if pd is not None and pd.notna(pgn_h_val):
                                            pgn_h_str = str(pgn_h_val).strip().upper()
                                            if pgn_h_str and pgn_h_str.lower() not in ['', 'nan', 'none', 'null', 'n/a', 'pgn(h)', 'pgn']:
                                                # Only count PGN for stats if this is a main message row
                                                if is_main_message_row:
                                                    all_pgn_h_values.append(pgn_h_str)
                                                    unique_pgn_h_values.add(pgn_h_str)
                                                # Also convert to decimal for PGN set
                                                try:
                                                    pgn_dec = int(pgn_h_str, 16)
                                                    if is_main_message_row:
                                                        pgns.add(pgn_dec)
                                                    current_pgn = pgn_dec
                                                    pgn_value = pgn_dec
                                                except ValueError:
                                                    pass
                                    except:
                                        pass
                            else:
                                if pgn_h_col_name and pgn_h_col_name in df:
                                    try:
                                        pgn_h_val = df[pgn_h_col_name][row_idx] if row_idx < len(df[pgn_h_col_name]) else None
                                        if pgn_h_val is not None:
                                            pgn_h_str = str(pgn_h_val).strip().upper()
                                            if pgn_h_str and pgn_h_str.lower() not in ['', 'nan', 'none', 'null', 'n/a', 'pgn(h)', 'pgn']:
                                                if is_main_message_row:
                                                    all_pgn_h_values.append(pgn_h_str)
                                                    unique_pgn_h_values.add(pgn_h_str)
                                                try:
                                                    pgn_dec = int(pgn_h_str, 16)
                                                    if is_main_message_row:
                                                        pgns.add(pgn_dec)
                                                    current_pgn = pgn_dec
                                                    pgn_value = pgn_dec
                                                except ValueError:
                                                    pass
                                    except:
                                        pass

                            # Get PGN (decimal column, fallback if no PGN(H))
                            if pgn_value is None:
                                if is_dataframe:
                                    if pgn_col_idx is not None:
                                        try:
                                            pgn_val = df.iloc[row_idx, pgn_col_idx]
                                            if pd is not None and pd.notna(pgn_val):
                                                pgn_value = int(float(str(pgn_val)))
                                                if 100 <= pgn_value <= 999999:
                                                    if is_main_message_row:
                                                        pgns.add(pgn_value)
                                                    current_pgn = pgn_value
                                            elif pgn_val is not None:
                                                pgn_value = int(float(str(pgn_val)))
                                                if 100 <= pgn_value <= 999999:
                                                    if is_main_message_row:
                                                        pgns.add(pgn_value)
                                                    current_pgn = pgn_value
                                        except:
                                            pass
                                else:
                                    if pgn_col_name and pgn_col_name in df:
                                        try:
                                            pgn_val = df[pgn_col_name][row_idx] if row_idx < len(df[pgn_col_name]) else None
                                            if pgn_val is not None:
                                                pgn_value = int(float(str(pgn_val)))
                                                if 100 <= pgn_value <= 999999:
                                                    if is_main_message_row:
                                                        pgns.add(pgn_value)
                                                    current_pgn = pgn_value
                                        except:
                                            pass

                            # Get SPN
                            if is_dataframe:
                                if spn_col_idx is not None:
                                    try:
                                        spn_val = df.iloc[row_idx, spn_col_idx]
                                        if pd is not None and pd.notna(spn_val):
                                            spn_value = int(float(str(spn_val)))
                                        elif spn_val is not None:
                                            spn_value = int(float(str(spn_val)))
                                    except:
                                        pass
                            else:
                                if spn_col_name and spn_col_name in df:
                                    try:
                                        spn_val = df[spn_col_name][row_idx] if row_idx < len(df[spn_col_name]) else None
                                        if spn_val is not None:
                                            spn_value = int(float(str(spn_val)))
                                    except:
                                        pass

                            # Get description
                            if is_dataframe:
                                if desc_col_idx is not None:
                                    try:
                                        desc_val = df.iloc[row_idx, desc_col_idx]
                                        if pd is not None and pd.notna(desc_val):
                                            desc_value = str(desc_val).strip()
                                        elif desc_val is not None:
                                            desc_value = str(desc_val).strip()
                                    except:
                                        pass
                            else:
                                if desc_col_name and desc_col_name in df:
                                    try:
                                        desc_val = df[desc_col_name][row_idx] if row_idx < len(df[desc_col_name]) else None
                                        if desc_val is not None:
                                            desc_value = str(desc_val).strip()
                                    except:
                                        pass

                            # Store SPN with PGN relationship
                            if spn_value is not None:
                                # Use PGN from current row, or fallback to last seen PGN
                                final_pgn = pgn_value if pgn_value else current_pgn
                                
                                if final_pgn:
                                    spns_data[(final_pgn, spn_value)] = desc_value
                                else:
                                    # Store without PGN if not found
                                    spns_data[(None, spn_value)] = desc_value

                        except Exception as row_exc:
                            logger.debug('Error processing row %d: %s', row_idx, str(row_exc))
                            continue

                # Fallback: extract from filename if vehicle name not found
                if not vehicle_name:
                    # Try to extract from filename (remove extension)
                    vehicle_name = os.path.splitext(fname)[0].strip()
                    if not vehicle_name or vehicle_name == '<unknown>':
                        vehicle_name = 'Unknown'

                # Default brand if not found
                if not brand:
                    brand = ''

                # Allow repeated uploads of the same vehicle on the same day (no duplicate blocking)

                # Get or create user
                try:
                    uploaded_by = request.user if getattr(request, 'user', None) and request.user.is_authenticated else None
                except Exception:
                    uploaded_by = None

                # Save Excel file for auditing (optional)
                excel_file_path = None
                try:
                    # Save file to media storage
                    file_path = f'j1939_uploads/{today.year}/{today.month:02d}/{today.day:02d}/{fname}'
                    saved_path = default_storage.save(file_path, f)
                    excel_file_path = saved_path
                    logger.info('Saved Excel file for auditing: %s', saved_path)
                except Exception as save_exc:
                    logger.warning('Failed to save Excel file for auditing: %s', str(save_exc))
                    # Continue processing even if file save fails

                # Create vehicle record
                vehicle = Vehicle.objects.create(
                    name=str(vehicle_name),
                    brand=str(brand),
                    uploaded_by=uploaded_by,
                    source_file=fname,
                    excel_file=excel_file_path if excel_file_path else None
                )

                # Create PGN records and associations
                vehicle_pgns = []
                for pgn_num in sorted(pgns):
                    pgn_obj, _ = PGN.objects.get_or_create(pgn_number=int(pgn_num))
                    VehiclePGN.objects.get_or_create(vehicle=vehicle, pgn=pgn_obj)
                    vehicle_pgns.append(pgn_num)

                # Create SPN records and associations with PGN relationships
                vehicle_spns = []
                for (pgn_num, spn_num), description in spns_data.items():
                    try:
                        pgn_obj = None
                        if pgn_num:
                            pgn_obj, _ = PGN.objects.get_or_create(pgn_number=int(pgn_num))

                        spn_obj, _ = SPN.objects.get_or_create(
                            spn_number=int(spn_num),
                            defaults={'description': description or ''}
                        )
                        # Update description if it was empty
                        if description and not spn_obj.description:
                            spn_obj.description = description
                            spn_obj.save()

                        vspn, created = VehicleSPN.objects.get_or_create(
                            vehicle=vehicle,
                            spn=spn_obj,
                            defaults={'pgn': pgn_obj}
                        )
                        if not created and pgn_obj and vspn.pgn != pgn_obj:
                            vspn.pgn = pgn_obj
                        if description:
                            vspn.value = description
                            vspn.supported = True
                        if pgn_obj and not vspn.pgn:
                            vspn.pgn = pgn_obj
                        vspn.save()

                        # Store for response (use PGN if available, otherwise None)
                        vehicle_spns.append({
                            'pgn': pgn_num if pgn_num else None,
                            'spn': spn_num,
                            'description': description or spn_obj.description or ''
                        })
                    except Exception as spn_exc:
                        logger.error('Error creating SPN %d: %s', spn_num, str(spn_exc))
                        continue

                # Map PGNs to SPNs from J1939ParameterDefinition table (J1939 Standard)
                j1939_mapped_spns = set()
                j1939_spn_details = []
                
                # Get all available PGNs from J1939ParameterDefinition for comparison
                available_pgns = set(J1939ParameterDefinition.objects.values_list('PGN_DEC', flat=True).distinct())
                matching_pgns = set(vehicle_pgns) & available_pgns
                logger.info('SPN Mapping: %d vehicle PGNs, %d in DB, %d matching', 
                           len(vehicle_pgns), len(available_pgns), len(matching_pgns))
                
                for pgn_num in vehicle_pgns:
                    # Get all SPNs defined for this PGN in the J1939 standard
                    spn_definitions = J1939ParameterDefinition.objects.filter(PGN_DEC=pgn_num)
                    for spn_def in spn_definitions:
                        j1939_mapped_spns.add(spn_def.SPN_Number)
                        j1939_spn_details.append({
                            'pgn': pgn_num,
                            'pgn_hex': spn_def.PGN_HEX,
                            'spn': spn_def.SPN_Number,
                            'description': spn_def.SPN_Description,
                            'unit': spn_def.Unit,
                            'resolution': float(spn_def.Resolution) if spn_def.Resolution else None,
                            'offset': float(spn_def.Offset) if spn_def.Offset else 0,
                            'start_byte': spn_def.Start_Byte,
                            'start_bit': spn_def.Start_Bit,
                            'bit_length': spn_def.Bit_Length,
                            'data_length_bytes': spn_def.Data_Length_Bytes
                        })
                
                logger.info('SPN Mapping Result: %d unique SPNs found from %d matching PGNs', 
                           len(j1939_mapped_spns), len(matching_pgns))

                # Build response data
                vehicles.append({
                    'id': vehicle.id,
                    'name': vehicle.name,
                    'brand': vehicle.brand,
                    'source_file': vehicle.source_file,
                    'pgns': vehicle_pgns,
                    'spns': vehicle_spns,
                    'pgn_count': len(vehicle_pgns),
                    'spn_count': len(vehicle_spns),
                    # PGN(H) column stats - calculated from pandas filtering by Index column
                    'total_pgn_messages': total_pgn_count,
                    'unique_pgn_count': unique_pgn_count,
                    'unique_pgn_list': unique_pgn_list,
                    # J1939 Standard SPN mapping (based on PGNs in file)
                    'j1939_unique_spn_count': len(j1939_mapped_spns),
                    'j1939_spn_list': sorted(list(j1939_mapped_spns)),
                    'j1939_spn_details': j1939_spn_details
                })

                logger.info('Successfully processed file %s: Vehicle=%s, Total PGN Messages=%d, Unique PGNs=%d',
                           fname, vehicle.name, total_pgn_count, unique_pgn_count)

            except Exception as exc:
                error_msg = f'Error processing file: {str(exc)}'
                errors.append({
                    'filename': fname,
                    'error': error_msg
                })
                logger.error('Exception processing %s: %s', fname, str(exc), exc_info=True)

        # Calculate aggregate totals across all vehicles
        total_pgn_messages_all = sum(v.get('total_pgn_messages', 0) for v in vehicles)
        all_unique_pgns = set()
        all_j1939_unique_spns = set()
        for v in vehicles:
            all_unique_pgns.update(v.get('unique_pgn_list', []))
            all_j1939_unique_spns.update(v.get('j1939_spn_list', []))

        # Build response - always return "success" status if request was processed
        # Errors are reported in the errors array
        response_data = {
            'status': 'success',
            'vehicles': vehicles,
            'errors': errors,
            'totals': {
                'total_vehicles': len(vehicles),
                'total_pgn_messages': total_pgn_messages_all,  # Sum of all PGN messages across all files
                'unique_pgn_count': len(all_unique_pgns),       # Total unique PGNs across all files
                'unique_pgn_list': sorted(list(all_unique_pgns)),
                # J1939 Standard SPN totals (mapped from PGNs)
                'j1939_unique_spn_count': len(all_j1939_unique_spns),
                'j1939_spn_list': sorted(list(all_j1939_unique_spns))
            }
        }

        # Return 200 OK even if there are errors, as long as the request was processed
        # Individual file errors are reported in the errors array
        status_code = status.HTTP_200_OK
        if not vehicles and errors:
            # If no vehicles were processed and there are errors, still return 200
            # but the errors array will contain the issues
            status_code = status.HTTP_200_OK

        return Response(response_data, status=status_code)


@method_decorator(csrf_exempt, name='dispatch')
class UploadAPIView(APIView):
    """
    New /api/upload/ endpoint per spec. Accepts a single file (or multiple) under 'file'.
    Parses Excel sheets, extracts vehicle name, brand, SPNs and PGNs and values, stores them, and
    returns a structured JSON summary per uploaded file.
    """
    permission_classes = [permissions.AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, format=None):
        files = request.FILES.getlist('file') or request.FILES.getlist('files[]') or list(request.FILES.values())
        if not files:
            return Response({'detail': 'No files uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        responses = []
        # header alias lists to support varied spreadsheets
        vehicle_name_aliases = ['vehicle name', 'veh name', 'vehicle', 'veh', 'unit name']
        brand_aliases = ['brand', 'make', 'manufacturer']
        spn_aliases = ['spn', 'spn number', 'spn_no', 'spn_number']
        pgn_aliases = ['pgn', 'pgn number', 'pgn_no', 'pgn_number']

        for f in files:
            fname = getattr(f, 'name', 'unknown')
            # validate extension
            if not (isinstance(fname, str) and (fname.lower().endswith('.xlsx') or fname.lower().endswith('.xls'))):
                responses.append({'filename': fname, 'error': 'Invalid file extension'})
                continue
            try:
                # Try openpyxl via pandas for robust sheet reading
                try:
                    xl = pd.read_excel(f, sheet_name=None, engine='openpyxl')
                except Exception:
                    # fallback: read via openpyxl directly
                    f.seek(0)
                    wb = load_workbook(filename=io.BytesIO(f.read()), data_only=True)
                    xl = {s.title: pd.DataFrame(wb[s].values) for s in wb.sheetnames}

                # Heuristic across sheets to find vehicle name/brand and SPNs/PGNs
                vehicle_name = None
                brand = None
                pgns = set()
                spns = {}  # spn_number -> value (string)

                for sheet_name, df in xl.items():
                    # Normalize DataFrame: if header row missing, create numeric columns
                    if isinstance(df.columns, pd.RangeIndex) and df.shape[0] > 0:
                        # attempt to check header-like first row
                        first_row = df.iloc[0].astype(str).fillna('').tolist()
                    else:
                        first_row = [str(x).strip().lower() for x in df.columns]

                    # Try to identify vehicle name / brand from header rows or first rows
                    joined_head = ' '.join([str(x) for x in first_row]).lower()
                    for alias in vehicle_name_aliases:
                        if alias in joined_head and not vehicle_name:
                            # attempt to locate column index
                            for i, val in enumerate(first_row):
                                if alias in str(val).lower():
                                    # look for non-empty cell in next rows
                                    for r in range(1, min(6, df.shape[0])):
                                        try:
                                            candidate = str(df.iloc[r, i])
                                        except Exception:
                                            candidate = ''
                                        if candidate and candidate.lower() not in ['nan', 'none', '']:
                                            vehicle_name = candidate.strip()
                                            break
                                    break
                    for alias in brand_aliases:
                        if alias in joined_head and not brand:
                            for i, val in enumerate(first_row):
                                if alias in str(val).lower():
                                    for r in range(1, min(6, df.shape[0])):
                                        try:
                                            candidate = str(df.iloc[r, i])
                                        except Exception:
                                            candidate = ''
                                        if candidate and candidate.lower() not in ['nan', 'none', '']:
                                            brand = candidate.strip()
                                            break
                                    break

                    # Detect SPN/PGN columns by header names
                    col_map = {i: str(c).strip().lower() for i, c in enumerate(df.columns)}
                    spn_cols = [i for i, h in col_map.items() if any(a in h for a in spn_aliases)]
                    pgn_cols = [i for i, h in col_map.items() if any(a in h for a in pgn_aliases)]

                    # If SPN/PGN columns found, use them
                    if spn_cols:
                        for i in spn_cols:
                            for r in range(1, df.shape[0]):
                                try:
                                    spn_val = df.iat[r, i]
                                except Exception:
                                    continue
                                if pd.isna(spn_val):
                                    continue
                                s = str(spn_val).strip()
                                # if like '190: 2200 RPM' parse
                                if ':' in s or '=' in s:
                                    parts = s.replace('=', ':').split(':')
                                    try:
                                        spnnum = int(parts[0])
                                        spns[spnnum] = parts[1].strip() if len(parts) > 1 else ''
                                    except Exception:
                                        # try to extract digits
                                        import re
                                        m = re.search(r"(\d{1,6})", s)
                                        if m:
                                            spnnum = int(m.group(1))
                                            remainder = s.replace(m.group(0), '').strip(' :\t')
                                            spns[spnnum] = remainder
                                else:
                                    try:
                                        num = int(s)
                                        spns[num] = ''
                                    except Exception:
                                        pass

                    if pgn_cols:
                        for i in pgn_cols:
                            for r in range(1, df.shape[0]):
                                try:
                                    pval = df.iat[r, i]
                                except Exception:
                                    continue
                                if pd.isna(pval):
                                    continue
                                try:
                                    valnum = int(str(pval).strip())
                                    if 100 <= valnum <= 999999:
                                        pgns.add(valnum)
                                except Exception:
                                    pass

                    # fallback scan: scan all cells for explicit SPN/PGN patterns
                    for col in df.columns:
                        try:
                            series = df[col].dropna().astype(str)
                        except Exception:
                            continue
                        for v in series:
                            vs = str(v)
                            low = vs.lower()
                            if 'spn' in low:
                                parts = vs.replace(':', ' ').replace('=', ' ').split()
                                for i_p, p in enumerate(parts):
                                    if p.lower().startswith('spn') and i_p + 1 < len(parts):
                                        try:
                                            spn_num = int(parts[i_p+1])
                                            val = None
                                            if i_p + 2 < len(parts):
                                                val = parts[i_p+2]
                                            spns[spn_num] = val or ''
                                        except Exception:
                                            pass
                            try:
                                num = int(vs)
                                if 100 <= num <= 999999:
                                    pgns.add(num)
                            except Exception:
                                pass

                # Ensure some defaults
                vehicle_name = vehicle_name or getattr(f, 'name', 'Unknown')
                brand = brand or ''

                uploaded_by = request.user if getattr(request, 'user', None) and request.user.is_authenticated else None

                vehicle = Vehicle.objects.create(name=str(vehicle_name), brand=str(brand), uploaded_by=uploaded_by, source_file=str(fname))

                # Persist PGNs
                for p in sorted(pgns):
                    pgn_obj, _ = PGN.objects.get_or_create(pgn_number=int(p))
                    VehiclePGN.objects.get_or_create(vehicle=vehicle, pgn=pgn_obj)

                # Persist SPNs
                for spn_num, val in spns.items():
                    spn_obj, _ = SPN.objects.get_or_create(spn_number=int(spn_num))
                    vspn, created = VehicleSPN.objects.get_or_create(vehicle=vehicle, spn=spn_obj)
                    # store value and mark supported if non-empty
                    if val is not None and str(val).strip() != '':
                        vspn.value = str(val)
                        vspn.supported = True
                        vspn.save()

                # compute summary
                total_pgns = VehiclePGN.objects.filter(vehicle=vehicle).count()
                total_spns = VehicleSPN.objects.filter(vehicle=vehicle).count()

                supported_spns_qs = VehicleSPN.objects.filter(vehicle=vehicle).select_related('spn')
                supported_spns = []
                for vs in supported_spns_qs:
                    supported_spns.append({'spn': vs.spn.spn_number, 'desc': vs.spn.description, 'value': vs.value if vs.supported else None, 'supported': bool(vs.supported)})

                responses.append({
                    'vehicle': vehicle.name,
                    'brand': vehicle.brand,
                    'spn_count': total_spns,
                    'pgn_count': total_pgns,
                    'source_file': vehicle.source_file,
                    'supported_spns': supported_spns,
                })

            except Exception as exc:
                tb = traceback.format_exc()
                logger.exception('Error processing uploaded file %s', fname)
                responses.append({'filename': fname, 'error': str(exc), 'traceback': tb})

        return Response({'results': responses}, status=status.HTTP_200_OK)


class VehicleListView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        vehicles = Vehicle.objects.select_related('uploaded_by').all().order_by('-upload_date')
        out = []
        for v in vehicles:
            pgn_count = VehiclePGN.objects.filter(vehicle=v).count()
            spn_count = VehicleSPN.objects.filter(vehicle=v).count()
            # Get uploader name
            uploaded_by_name = None
            if v.uploaded_by:
                uploaded_by_name = f"{v.uploaded_by.first_name} {v.uploaded_by.last_name}".strip()
                if not uploaded_by_name:
                    uploaded_by_name = v.uploaded_by.username
            out.append({
                'id': v.id, 
                'name': v.name, 
                'brand': v.brand, 
                'source_file': v.source_file,
                'pgns': pgn_count, 
                'spns': spn_count,
                'uploaded_by': v.uploaded_by.id if v.uploaded_by else None,
                'uploaded_by_name': uploaded_by_name,
                'upload_date': v.upload_date.isoformat() if v.upload_date else None
            })
        return Response(out)


class VehicleSpnsView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request, vehicle_id):
        try:
            vehicle = Vehicle.objects.get(pk=vehicle_id)
        except Vehicle.DoesNotExist:
            return Response({'detail': 'Vehicle not found'}, status=status.HTTP_404_NOT_FOUND)

        spns = VehicleSPN.objects.filter(vehicle=vehicle).select_related('spn', 'pgn')
        spn_payload = [
            {
                'pgn': s.pgn.pgn_number if s.pgn else None,
                'spn': s.spn.spn_number,
                'description': s.spn.description,
                'supported': s.supported,
                'value': s.value if s.supported else None
            }
            for s in spns
        ]
        return Response({
            'vehicle_id': vehicle.id,
            'vehicle': vehicle.name,
            'brand': vehicle.brand,
            'spns': spn_payload
        })


class SpnVehiclesView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request, spn_number):
        try:
            spn_number = int(spn_number)
        except (TypeError, ValueError):
            return Response({'detail': 'Invalid SPN number'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            spn = SPN.objects.get(spn_number=spn_number)
            description = spn.description
        except SPN.DoesNotExist:
            description = ''
            spn = None

        vehicles = []
        if spn:
            vspns = VehicleSPN.objects.filter(spn=spn, supported=True).select_related('vehicle')
            vehicles = [
                {
                    'name': v.vehicle.name,
                    'brand': v.vehicle.brand
                }
                for v in vspns
            ]

        return Response({
            'spn': spn_number,
            'description': description,
            'vehicles': vehicles
        })


# Standard Files and Auxiliary Files Views
class StandardFileListView(generics.ListCreateAPIView):
    """List and create standard files"""
    queryset = StandardFile.objects.all()
    serializer_class = StandardFileSerializer
    permission_classes = [permissions.AllowAny]  # Allow unauthenticated access for now
    parser_classes = [MultiPartParser, FormParser]

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user if self.request.user.is_authenticated else None)


class StandardFileDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update, delete standard files"""
    queryset = StandardFile.objects.all()
    serializer_class = StandardFileSerializer
    permission_classes = [permissions.AllowAny]  # Allow unauthenticated access for now
    parser_classes = [MultiPartParser, FormParser]


class AuxiliaryFileListView(generics.ListCreateAPIView):
    """List and create auxiliary files"""
    queryset = AuxiliaryFile.objects.all()
    serializer_class = AuxiliaryFileSerializer
    permission_classes = [permissions.AllowAny]  # Allow unauthenticated access for now
    parser_classes = [MultiPartParser, FormParser]

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user if self.request.user.is_authenticated else None)


class AuxiliaryFileDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update, delete auxiliary files"""
    queryset = AuxiliaryFile.objects.all()
    serializer_class = AuxiliaryFileSerializer
    permission_classes = [permissions.AllowAny]  # Allow unauthenticated access for now
    parser_classes = [MultiPartParser, FormParser]


class CategoryListView(generics.ListCreateAPIView):
    """List and create categories"""
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.AllowAny]  # Allow unauthenticated access for now

    def perform_create(self, serializer):
        serializer.save()


class CategoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update, delete categories"""
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.AllowAny]  # Allow unauthenticated access for now


class PGNListView(generics.ListAPIView):
    """List all PGNs"""
    queryset = PGN.objects.all()
    serializer_class = PGNSerializer
    permission_classes = [permissions.AllowAny]


class SPNListView(generics.ListAPIView):
    """List all SPNs"""
    queryset = SPN.objects.all()
    serializer_class = SPNSerializer
    permission_classes = [permissions.AllowAny]


# ---------------------------------------------------------------------------
# Helpers: J1939 PGN→SPN map loading and summarization
# ---------------------------------------------------------------------------
_J1939_MAP_CACHE = None


def load_j1939_map():
    """
    Load J1939 PGN→SPN map from Main/j1939_map.json if available.
    Expected shape:
    {
      "61444": {
        "pgn_name": "...",
        "spns": [ {"spn": 190, "name": "Engine Speed"}, ... ]
      },
      "F004": { ... }  # hex keys also allowed
    }
    """
    global _J1939_MAP_CACHE
    if _J1939_MAP_CACHE is not None:
        return _J1939_MAP_CACHE
    try:
        map_path = os.path.join(os.path.dirname(__file__), "j1939_map.json")
        with open(map_path, "r", encoding="utf-8") as fh:
            _J1939_MAP_CACHE = json.load(fh)
    except Exception:
        _J1939_MAP_CACHE = {}
    return _J1939_MAP_CACHE


def summarize_pgns_with_map(pgn_list, j1939_map):
    """
    Given a list of PGN dicts (with pgn_hex/pgn_dec) and a J1939 map,
    return per-PGN SPN mapping and counts.
    """
    if not j1939_map:
        return {
            "pgn_count": 0,
            "spn_occurrences": 0,
            "unique_spn_count": 0,
            "pgn_spn_mapping": {}
        }

    pgn_spn_mapping = {}
    unique_spns = set()
    total_spn_occurrences = 0

    for p in pgn_list:
        pgn_hex = str(p.get('pgn_hex') or '').upper()
        pgn_dec = p.get('pgn_dec')

        # Try matching by hex then dec string
        entry = None
        if pgn_hex and pgn_hex in j1939_map:
            entry = j1939_map[pgn_hex]
        elif pgn_dec is not None and str(pgn_dec) in j1939_map:
            entry = j1939_map[str(pgn_dec)]

        if not entry:
            continue

        spns = entry.get("spns", [])
        spn_items = []
        for s in spns:
            spn_num = s.get("spn")
            if spn_num in (None, 0):
                continue
            unique_spns.add(spn_num)
            total_spn_occurrences += 1
            spn_items.append({
                "spn": spn_num,
                "name": s.get("name", f"SPN_{spn_num}")
            })

        pgn_spn_mapping[str(p.get('pgn_dec') or pgn_hex)] = {
            "pgn": p.get('pgn_dec') if p.get('pgn_dec') else pgn_hex,
            "name": entry.get("pgn_name", p.get("name", "")),
            "spns": spn_items,
            "spn_count": len(spn_items)
        }

    return {
        "pgn_count": len(pgn_spn_mapping),
        "spn_occurrences": total_spn_occurrences,
        "unique_spn_count": len(unique_spns),
        "pgn_spn_mapping": pgn_spn_mapping
    }


# ---------------------------------------------------------------------------
# J1939 Data Log File Analysis with Robust Encoding Detection
# Parses uploaded CSVs, maps PGNs/SPNs, and returns aggregates.
# ---------------------------------------------------------------------------

def detect_encoding_and_read(file_content):
    """
    Detect encoding and read file content.
    Try multiple encodings in order of likelihood.
    Returns tuple: (decoded_text, encoding_used, errors_list)
    """
    encodings_to_try = ['utf-8', 'utf-8-sig', 'utf-16', 'utf-16-le', 'utf-16-be', 
                        'gb2312', 'gbk', 'gb18030', 'big5', 'latin1', 'cp1252', 
                        'iso-8859-1', 'ascii']
    
    errors_list = []
    
    for encoding in encodings_to_try:
        try:
            decoded = file_content.decode(encoding)
            # Verify it's readable text (has reasonable characters)
            if decoded and len(decoded) > 0:
                # Check if it contains mostly printable characters
                printable_ratio = sum(1 for c in decoded[:1000] if c.isprintable() or c in '\n\r\t') / min(len(decoded), 1000)
                if printable_ratio > 0.7:  # At least 70% printable
                    return decoded, encoding, []
        except (UnicodeDecodeError, UnicodeError) as e:
            errors_list.append(f"{encoding}: {str(e)[:50]}")
            continue
        except Exception as e:
            errors_list.append(f"{encoding}: {str(e)[:50]}")
            continue
    
    # Fallback: binary-safe reading with error replacement
    try:
        decoded = file_content.decode('utf-8', errors='replace')
        return decoded, 'utf-8-fallback', errors_list
    except Exception as e:
        # Last resort: latin1 always works for any byte sequence
        decoded = file_content.decode('latin1', errors='replace')
        return decoded, 'latin1-fallback', errors_list


def extract_pgn_from_can_id(can_id_value):
    """
    Extract PGN from CAN ID according to J1939 specification.
    PGN = (CAN_ID >> 8) & 0x3FFFF for extended frames
    
    For standard J1939:
    - 29-bit CAN ID: Priority (3 bits) + Reserved (1 bit) + Data Page (1 bit) + PDU Format (8 bits) + 
                     PDU Specific/Destination (8 bits) + Source Address (8 bits)
    - PGN = bits 8-25 (18 bits max), but commonly 16 bits: (CAN_ID >> 8) & 0xFFFF
    
    Args:
        can_id_value: CAN ID as int, hex string, or decimal string
        
    Returns:
        tuple: (pgn_int, pgn_hex_str) or (None, None) if invalid
    """
    try:
        # Convert to integer
        if isinstance(can_id_value, int):
            can_id = can_id_value
        elif isinstance(can_id_value, str):
            can_id_str = can_id_value.strip().upper()
            # Remove common prefixes
            for prefix in ['0X', '0H', 'H', 'X']:
                if can_id_str.startswith(prefix):
                    can_id_str = can_id_str[len(prefix):]
            
            if not can_id_str:
                return None, None
            
            # Try hex first if it looks like hex
            if re.match(r'^[0-9A-F]+$', can_id_str):
                # Determine if it's hex or decimal
                # If contains A-F, definitely hex
                if re.search(r'[A-F]', can_id_str):
                    can_id = int(can_id_str, 16)
                # If length > 8 digits, likely hex
                elif len(can_id_str) > 8:
                    can_id = int(can_id_str, 16)
                # Try as decimal first for pure numeric
                else:
                    try:
                        can_id = int(can_id_str, 10)
                        # If result is unreasonably large, try hex
                        if can_id > 0x1FFFFFFF:  # Max 29-bit CAN ID
                            can_id = int(can_id_str, 16)
                    except ValueError:
                        can_id = int(can_id_str, 16)
            else:
                # Pure decimal
                can_id = int(can_id_str, 10)
        else:
            return None, None
        
        # Validate CAN ID range (29-bit max)
        if can_id < 0 or can_id > 0x1FFFFFFF:
            return None, None
        
        # Extract PGN: (CAN_ID >> 8) & 0xFFFF (16-bit PGN)
        # For J1939, PGN is typically in bits 8-23
        pgn = (can_id >> 8) & 0xFFFF
        
        # Format as hex string (uppercase, no prefix)
        pgn_hex = f"{pgn:04X}"
        
        return pgn, pgn_hex
        
    except (ValueError, TypeError) as e:
        return None, None


def find_can_id_column(headers):
    """
    Find the column index that likely contains CAN ID.
    Returns column index or None.
    """
    can_id_patterns = [
        r'^can\s*id$', r'^canid$', r'^id$', r'^can_id$',
        r'^message\s*id$', r'^msg\s*id$', r'^msgid$',
        r'^arbitration\s*id$', r'^arb\s*id$',
        r'^identifier$', r'^frame\s*id$',
        r'^id\(h\)$', r'^id_h$', r'^id\s*\(hex\)$'
    ]
    
    for idx, header in enumerate(headers):
        header_clean = header.strip().lower()
        for pattern in can_id_patterns:
            if re.match(pattern, header_clean):
                return idx
    return None


def find_pgn_column(headers):
    """
    Find the column index that contains PGN values directly.
    Returns column index or None.
    """
    pgn_patterns = [
        r'^pgn\s*\(h\)$', r'^pgn_h$', r'^pgn\s*\(hex\)$', r'^pgn\s*hex$',
        r'^pgn$', r'^pgn_dec$', r'^pgn\s*\(d\)$', r'^pgn\s*\(dec\)$'
    ]
    
    for idx, header in enumerate(headers):
        header_clean = header.strip().lower()
        for pattern in pgn_patterns:
            if re.match(pattern, header_clean):
                return idx
    return None


def parse_line_for_can_id(line, delimiter=None):
    """
    Parse a line and try to extract CAN ID from various formats.
    Handles CSV, space-separated, and raw CAN frame formats.
    
    Returns list of potential CAN IDs found.
    """
    can_ids = []
    
    # Try different delimiters
    delimiters_to_try = [delimiter] if delimiter else [',', ';', '\t', ' ', '|']
    
    for delim in delimiters_to_try:
        if delim is None:
            continue
        parts = [p.strip() for p in line.split(delim) if p.strip()]
        
        for part in parts:
            # Check if part looks like a CAN ID (hex or decimal)
            part_clean = part.upper().strip()
            
            # Remove common prefixes
            for prefix in ['0X', '0H', 'H']:
                if part_clean.startswith(prefix):
                    part_clean = part_clean[len(prefix):]
            
            # Check if it's a valid hex/decimal number in CAN ID range
            if re.match(r'^[0-9A-F]+$', part_clean):
                try:
                    # Try as hex first
                    if re.search(r'[A-F]', part_clean):
                        val = int(part_clean, 16)
                    else:
                        # Pure numeric - could be decimal or hex
                        val = int(part_clean, 10)
                        if val > 0x1FFFFFFF:
                            val = int(part_clean, 16)
                    
                    # CAN ID reasonable range (J1939 29-bit extended)
                    if 0x100 <= val <= 0x1FFFFFFF:
                        can_ids.append(val)
                except ValueError:
                    continue
    
    return can_ids


@csrf_exempt
@require_POST
def analyze_j1939_files(request):
    """
    Analyze J1939 data log files and extract PGNs.
    
    Supports:
    - .csv, .txt, .log, and binary log files
    - Automatic encoding detection (UTF-8, UTF-16, latin1, cp1252, GB2312, etc.)
    - PGN extraction from CAN ID: PGN = (ID >> 8) & 0xFFFF
    - Handles CAN ID in hex or decimal format
    - CSV, space-separated, and raw frame formats
    - Large file handling with streaming
    - Graceful error handling for encoding issues
    
    Returns JSON:
    {
        "status": "success",
        "total_pgn_count": <int>,      # Total PGN occurrences
        "unique_pgn_count": <int>,      # Unique PGN values
        "unique_pgn_list": ["F004", "FEF2", ...],
        "vehicles": [...],              # Per-file breakdown
        "errors": [...]                 # Any parsing errors
    }
    """
    files = request.FILES.getlist('files')
    if not files:
        files = request.FILES.getlist('file')
    if not files:
        return JsonResponse({
            'status': 'error',
            'message': 'No files uploaded',
            'total_pgn_count': 0,
            'unique_pgn_count': 0,
            'unique_pgn_list': []
        }, status=400)
    
    vehicles = []
    errors = []
    
    # Aggregate counters across all files
    all_pgn_occurrences = []  # List of all PGN hex values (with duplicates)
    all_unique_pgns = set()   # Set of unique PGN hex values
    
    for file in files:
        file_errors = []
        file_pgn_occurrences = []
        file_unique_pgns = set()
        
        try:
            # Read file content
            file_content = file.read()
            file.seek(0)  # Reset for potential re-read
            
            # Detect encoding and decode
            decoded_text, encoding_used, encoding_errors = detect_encoding_and_read(file_content)
            
            if encoding_errors:
                file_errors.append(f"Encoding detection tried: {', '.join(encoding_errors[:3])}")
            
            logger.info(f"File {file.name}: Using encoding {encoding_used}")
            
            # Split into lines
            lines = decoded_text.split('\n')
            
            # Detect file format and find relevant columns
            headers = []
            pgn_col_idx = None
            can_id_col_idx = None
            delimiter = ','
            
            # Find header row and column indices
            for i, line in enumerate(lines[:20]):  # Check first 20 lines for header
                line = line.strip()
                if not line:
                    continue
                
                # Detect delimiter
                for delim in [',', ';', '\t', '|']:
                    if delim in line:
                        delimiter = delim
                        break
                
                columns = [col.strip() for col in line.split(delimiter)]
                
                # Check if this looks like a header row
                pgn_idx = find_pgn_column(columns)
                can_idx = find_can_id_column(columns)
                
                if pgn_idx is not None or can_idx is not None:
                    headers = columns
                    pgn_col_idx = pgn_idx
                    can_id_col_idx = can_idx
                    # Start processing from next line
                    lines = lines[i+1:]
                    break
            
            # Process data lines
            for line_num, line in enumerate(lines):
                try:
                    line = line.strip()
                    if not line:
                        continue
                    
                    columns = [col.strip() for col in line.split(delimiter)]
                    
                    # Method 1: Direct PGN column
                    if pgn_col_idx is not None and pgn_col_idx < len(columns):
                        pgn_value = columns[pgn_col_idx].strip()
                        if pgn_value and pgn_value.lower() not in ['', 'null', 'none', 'n/a', 'pgn', 'pgn(h)']:
                            # Normalize PGN value to hex
                            pgn_clean = pgn_value.upper()
                            for prefix in ['0X', '0H', 'H']:
                                if pgn_clean.startswith(prefix):
                                    pgn_clean = pgn_clean[len(prefix):]
                            
                            if re.match(r'^[0-9A-F]+$', pgn_clean):
                                # Pad to at least 4 characters
                                pgn_hex = pgn_clean.zfill(4).upper()
                                file_pgn_occurrences.append(pgn_hex)
                                file_unique_pgns.add(pgn_hex)
                    
                    # Method 2: Extract PGN from CAN ID column
                    elif can_id_col_idx is not None and can_id_col_idx < len(columns):
                        can_id_value = columns[can_id_col_idx].strip()
                        pgn_int, pgn_hex = extract_pgn_from_can_id(can_id_value)
                        if pgn_hex:
                            file_pgn_occurrences.append(pgn_hex)
                            file_unique_pgns.add(pgn_hex)
                    
                    # Method 3: Try to find CAN ID anywhere in the line
                    else:
                        can_ids = parse_line_for_can_id(line, delimiter)
                        for can_id in can_ids[:1]:  # Take first valid CAN ID per line
                            pgn_int, pgn_hex = extract_pgn_from_can_id(can_id)
                            if pgn_hex:
                                file_pgn_occurrences.append(pgn_hex)
                                file_unique_pgns.add(pgn_hex)
                                break
                
                except Exception as line_error:
                    # Log but continue processing
                    if len(file_errors) < 10:
                        file_errors.append(f"Line {line_num}: {str(line_error)[:50]}")
                    continue
            
            # Build vehicle/file result
            vehicle_name = file.name.split('.')[0].replace('_', ' ').replace('-', ' ')
            
            vehicle = {
                'id': len(vehicles) + 1,
                'name': vehicle_name,
                'brand': extract_brand(file.name),
                'filename': file.name,
                'encoding_used': encoding_used,
                'total_pgn_count': len(file_pgn_occurrences),
                'unique_pgn_count': len(file_unique_pgns),
                'unique_pgn_list': sorted(list(file_unique_pgns)),
                'analysis_summary': {
                    'total_lines_processed': len(lines),
                    'pgn_extraction_method': 'pgn_column' if pgn_col_idx is not None else 
                                            'can_id_column' if can_id_col_idx is not None else 
                                            'auto_detect'
                }
            }
            
            vehicles.append(vehicle)
            
            # Add to aggregates
            all_pgn_occurrences.extend(file_pgn_occurrences)
            all_unique_pgns.update(file_unique_pgns)
            
            if file_errors:
                errors.append({
                    'filename': file.name,
                    'warnings': file_errors[:10]  # Limit warnings per file
                })
                
        except Exception as e:
            logger.error(f"Error processing file {file.name}: {str(e)}", exc_info=True)
            errors.append({
                'filename': file.name,
                'error': f"Failed to parse file: {str(e)}"
            })
            # Still add empty vehicle entry
            vehicles.append({
                'id': len(vehicles) + 1,
                'name': file.name.split('.')[0],
                'filename': file.name,
                'total_pgn_count': 0,
                'unique_pgn_count': 0,
                'unique_pgn_list': [],
                'error': str(e)
            })
    
    # Build response
    return JsonResponse({
        'status': 'success',
        'total_pgn_count': len(all_pgn_occurrences),
        'unique_pgn_count': len(all_unique_pgns),
        'unique_pgn_list': sorted(list(all_unique_pgns)),
        'vehicles': vehicles,
        'errors': errors,
        'totals': {
            'total_vehicles': len(vehicles),
            'total_pgn_count': len(all_pgn_occurrences),
            'unique_pgn_count': len(all_unique_pgns),
            'pgn_h_column_stats': {
                'total_pgn_count': len(all_pgn_occurrences),
                'unique_pgn_count': len(all_unique_pgns)
            }
        }
    })


def extract_brand(filename):
    """Extract vehicle brand from filename"""
    filename_lower = filename.lower()
    brands = {
        'daf': 'DAF',
        'hino': 'HINO',
        'volvo': 'Volvo',
        'scania': 'Scania',
        'mercedes': 'Mercedes',
        'man': 'MAN',
        'iveco': 'Iveco',
        'kenworth': 'Kenworth',
        'peterbilt': 'Peterbilt'
    }
    for brand_key, brand_name in brands.items():
        if brand_key in filename_lower:
            return brand_name
    return 'Unknown'


# =============================================================================
# J1939 Parameter Definitions API Views
# =============================================================================

class J1939ParameterDefinitionListView(generics.ListCreateAPIView):
    """
    GET: List all J1939 parameter definitions
    POST: Create a new parameter definition
    """
    queryset = J1939ParameterDefinition.objects.all()
    serializer_class = J1939ParameterDefinitionSerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        queryset = J1939ParameterDefinition.objects.all()
        # Filter by PGN if provided
        pgn = self.request.query_params.get('pgn', None)
        if pgn is not None:
            queryset = queryset.filter(PGN_DEC=pgn)
        return queryset


class J1939ParameterDefinitionDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    GET: Retrieve a specific parameter definition by SPN number
    PUT/PATCH: Update a parameter definition
    DELETE: Delete a parameter definition
    """
    queryset = J1939ParameterDefinition.objects.all()
    serializer_class = J1939ParameterDefinitionSerializer
    permission_classes = [permissions.AllowAny]
    lookup_field = 'SPN_Number'


class DecodeSPNValueView(APIView):
    """
    POST /api/j1939/decode-spn/
    
    Decode raw CAN data bytes to physical SPN value using J1939 parameter definitions.
    
    Request Body:
    {
        "spn_number": 84,
        "raw_data": [0, 0, 0, 0, 255, 127, 0, 0]  // 8 bytes of raw CAN data
    }
    
    Response:
    {
        "spn_number": 84,
        "spn_description": "Wheel-Based Vehicle Speed",
        "raw_value": 32767,
        "physical_value": 255.9921875,
        "unit": "km/h",
        "status": "Valid"
    }
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = SPNDecodeRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        spn_number = serializer.validated_data['spn_number']
        raw_data = serializer.validated_data['raw_data']

        try:
            param_def = J1939ParameterDefinition.objects.get(SPN_Number=spn_number)
        except J1939ParameterDefinition.DoesNotExist:
            return Response({
                'error': f'SPN {spn_number} not found in parameter definitions'
            }, status=status.HTTP_404_NOT_FOUND)

        # Decode the SPN value
        decoded = param_def.decode_spn_value(raw_data)

        if decoded is None:
            return Response({
                'spn_number': spn_number,
                'spn_description': param_def.SPN_Description,
                'raw_value': None,
                'physical_value': None,
                'unit': param_def.Unit,
                'status': 'Error - Could not decode raw bytes'
            })

        response_data = {
            'spn_number': decoded['spn_number'],
            'spn_description': decoded['spn_description'],
            'raw_value': decoded['raw_value'],
            'physical_value': decoded['physical_value'],
            'unit': decoded['unit'],
            'status': decoded['status']
        }

        return Response(response_data)


class UniqueSPNCountView(APIView):
    """
    GET /api/j1939/unique-spn-count/
    
    Get count of unique SPNs in the J1939 parameter definitions.
    Optionally filter by PGN.
    
    Query Parameters:
    - pgn (optional): Filter by PGN decimal number
    
    Response:
    {
        "unique_spn_count": 9,
        "spn_numbers": [84, 182, 959, 960, 961, 962, 963, 964, 4191],
        "pgn_filter": null
    }
    """
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        pgn = request.query_params.get('pgn', None)
        
        queryset = J1939ParameterDefinition.objects.all()
        if pgn is not None:
            try:
                pgn = int(pgn)
                queryset = queryset.filter(PGN_DEC=pgn)
            except ValueError:
                return Response({
                    'error': 'Invalid PGN value'
                }, status=status.HTTP_400_BAD_REQUEST)

        spn_numbers = list(queryset.values_list('SPN_Number', flat=True).distinct())
        
        return Response({
            'unique_spn_count': len(spn_numbers),
            'spn_numbers': sorted(spn_numbers),
            'pgn_filter': pgn
        })


class PGNSummaryView(APIView):
    """
    GET /api/j1939/pgn-summary/
    
    Get summary of all PGNs with their associated SPNs from parameter definitions.
    
    Response:
    {
        "pgn_count": 4,
        "pgns": [
            {
                "pgn": 0,
                "spn_count": 1,
                "spns": [4191]
            },
            ...
        ]
    }
    """
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        # Group SPNs by PGN
        pgn_data = J1939ParameterDefinition.objects.values('PGN_DEC').annotate(
            spn_count=Count('SPN_Number')
        ).order_by('PGN_DEC')

        pgns = []
        for item in pgn_data:
            spns = list(J1939ParameterDefinition.objects.filter(
                PGN_DEC=item['PGN_DEC']
            ).values_list('SPN_Number', flat=True))
            pgns.append({
                'pgn': item['PGN_DEC'],
                'spn_count': item['spn_count'],
                'spns': sorted(spns)
            })

        return Response({
            'pgn_count': len(pgns),
            'pgns': pgns
        })


class PGNToSPNMappingView(APIView):
    """
    POST /api/j1939/pgn-spn-mapping/
    
    Map a list of PGNs to their corresponding SPNs with full details.
    This is the main endpoint for SPN analysis from uploaded files.
    
    Request Body:
    {
        "pgn_list": [65265, 65254, 65263, 65266, 61450, ...]
    }
    
    Response:
    {
        "total_unique_spn_count": 11,
        "total_pgn_count": 5,
        "pgns_with_spn_data": 4,
        "pgns_without_spn_data": 1,
        "data": [
            {
                "PGN_DEC": 65266,
                "PGN_HEX": "0xFEF2",
                "SPNs": [
                    {
                        "SPN": 183,
                        "Description": "Engine Fuel Rate",
                        "Bit_Range": "1-2",
                        "Length_Bits": 16,
                        "Unit": "L/h",
                        "Resolution": 0.05,
                        "Offset": 0
                    }
                ]
            },
            {
                "PGN_DEC": 61450,
                "PGN_HEX": "0xF00A",
                "SPNs": [],
                "message": "No SPN Data Available"
            }
        ]
    }
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        pgn_list = request.data.get('pgn_list', [])
        
        if not pgn_list:
            return Response({
                'error': 'pgn_list is required',
                'message': 'Please provide a list of PGN decimal values'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Convert all to integers
        try:
            pgn_list = [int(pgn) for pgn in pgn_list]
        except (ValueError, TypeError):
            return Response({
                'error': 'Invalid PGN values',
                'message': 'All PGN values must be integers'
            }, status=status.HTTP_400_BAD_REQUEST)

        result = []
        unique_spns = set()
        pgns_with_data = 0
        pgns_without_data = 0

        for pgn in pgn_list:
            # Get all SPN definitions for this PGN
            spn_rows = J1939ParameterDefinition.objects.filter(PGN_DEC=pgn)

            if not spn_rows.exists():
                pgns_without_data += 1
                result.append({
                    "PGN_DEC": pgn,
                    "PGN_HEX": f"0x{pgn:04X}",
                    "SPNs": [],
                    "message": "No SPN Data Available"
                })
                continue

            pgns_with_data += 1
            spn_list = []
            pgn_hex = None

            for row in spn_rows:
                unique_spns.add(row.SPN_Number)
                pgn_hex = row.PGN_HEX or f"0x{pgn:04X}"

                # Calculate bit range string
                start_byte = row.Start_Byte
                bit_length = row.Bit_Length
                data_length = row.Data_Length_Bytes
                
                if data_length == 1:
                    bit_range = str(start_byte)
                elif data_length == 2:
                    bit_range = f"{start_byte}-{start_byte + 1}"
                else:
                    bit_range = f"{start_byte}-{start_byte + data_length - 1}"

                spn_list.append({
                    "SPN": row.SPN_Number,
                    "Description": row.SPN_Description,
                    "Bit_Range": bit_range,
                    "Start_Byte": start_byte,
                    "Start_Bit": row.Start_Bit,
                    "Length_Bits": bit_length,
                    "Data_Length_Bytes": data_length,
                    "Unit": row.Unit,
                    "Resolution": row.Resolution,
                    "Offset": row.Offset,
                    "Min_Value": row.Min_Value,
                    "Max_Value": row.Max_Value
                })

            result.append({
                "PGN_DEC": pgn,
                "PGN_HEX": pgn_hex,
                "SPNs": spn_list
            })

        return Response({
            "total_unique_spn_count": len(unique_spns),
            "total_pgn_count": len(pgn_list),
            "pgns_with_spn_data": pgns_with_data,
            "pgns_without_spn_data": pgns_without_data,
            "unique_spn_list": sorted(list(unique_spns)),
            "data": result
        })


class UploadSPNMasterView(APIView):
    """
    POST /api/j1939/upload-spn-master/
    
    Upload a CSV file containing SPN master data to populate the J1939ParameterDefinition table.
    
    Expected CSV columns:
    PGN_DEC, PGN_HEX, SPN_Number, SPN_Name, DL, SPB, Length_Bits, Unit
    
    Optional columns:
    Resolution, Offset, Min_Value, Max_Value
    """
    permission_classes = [permissions.AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if 'file' not in request.FILES:
            return Response({
                'error': 'No file uploaded',
                'message': 'Please upload a CSV file'
            }, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        
        # Check file extension
        if not file.name.endswith('.csv'):
            return Response({
                'error': 'Invalid file type',
                'message': 'Only CSV files are supported'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Read CSV file
            import csv
            import io
            
            content = file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(content))
            
            created_count = 0
            updated_count = 0
            errors = []

            for row_num, row in enumerate(reader, start=2):
                try:
                    # Parse required fields
                    spn_number = int(row.get('SPN_Number', row.get('SPN', 0)))
                    pgn_dec = int(row.get('PGN_DEC', row.get('PGN', 0)))
                    pgn_hex = row.get('PGN_HEX', f"0x{pgn_dec:04X}")
                    spn_name = row.get('SPN_Name', row.get('Description', ''))
                    unit = row.get('Unit', '')
                    
                    # Parse byte/bit info
                    data_length = int(row.get('DL', row.get('Data_Length', 1)))
                    spb = row.get('SPB', row.get('Start_Byte', '1'))
                    
                    # Parse SPB which could be "1-2" or just "1"
                    if '-' in str(spb):
                        start_byte = int(str(spb).split('-')[0])
                    else:
                        start_byte = int(spb)
                    
                    length_bits = int(row.get('Length_Bits', row.get('Bit_Length', 8)))
                    
                    # Parse optional fields
                    resolution = float(row.get('Resolution', 1.0))
                    offset = float(row.get('Offset', 0.0))
                    min_value = row.get('Min_Value')
                    max_value = row.get('Max_Value')
                    
                    min_value = float(min_value) if min_value else None
                    max_value = float(max_value) if max_value else None

                    # Create or update
                    obj, created = J1939ParameterDefinition.objects.update_or_create(
                        SPN_Number=spn_number,
                        defaults={
                            'PGN_DEC': pgn_dec,
                            'PGN_HEX': pgn_hex,
                            'SPN_Description': spn_name,
                            'Unit': unit,
                            'Data_Length_Bytes': data_length,
                            'Start_Byte': start_byte,
                            'Start_Bit': 0,
                            'Bit_Length': length_bits,
                            'Resolution': resolution,
                            'Offset': offset,
                            'Min_Value': min_value,
                            'Max_Value': max_value
                        }
                    )

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")

            return Response({
                'status': 'success',
                'created': created_count,
                'updated': updated_count,
                'total_processed': created_count + updated_count,
                'errors': errors if errors else None,
                'message': f"Successfully processed {created_count + updated_count} SPN definitions"
            })

        except Exception as e:
            return Response({
                'error': 'Failed to process file',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AnalyzePGNsFromFileView(APIView):
    """
    POST /api/j1939/analyze-pgns/
    
    Analyze PGNs from an uploaded J1939 data file and return SPN mappings.
    Automatically extracts PGNs from the file and maps them to SPNs.
    """
    permission_classes = [permissions.AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if 'file' not in request.FILES:
            return Response({
                'error': 'No file uploaded'
            }, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        extracted_pgns = set()
        
        try:
            content = file.read().decode('utf-8', errors='ignore')
            
            # Try to parse as CSV
            import csv
            import io
            
            reader = csv.DictReader(io.StringIO(content))
            
            # Look for PGN column (various possible names)
            pgn_columns = ['PGN', 'pgn', 'PGN_DEC', 'pgn_dec', 'PGN_H', 'PGN_Hex', 'ParameterGroupNumber']
            
            for row in reader:
                for col in pgn_columns:
                    if col in row and row[col]:
                        try:
                            pgn_value = row[col]
                            # Handle hex values
                            if str(pgn_value).lower().startswith('0x'):
                                pgn_int = int(pgn_value, 16)
                            else:
                                pgn_int = int(pgn_value)
                            extracted_pgns.add(pgn_int)
                        except (ValueError, TypeError):
                            pass

            if not extracted_pgns:
                return Response({
                    'error': 'No PGNs found in file',
                    'message': 'Could not extract PGN values from the uploaded file'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Now map PGNs to SPNs
            pgn_list = sorted(list(extracted_pgns))
            
            result = []
            unique_spns = set()
            pgns_with_data = 0
            pgns_without_data = 0

            for pgn in pgn_list:
                spn_rows = J1939ParameterDefinition.objects.filter(PGN_DEC=pgn)

                if not spn_rows.exists():
                    pgns_without_data += 1
                    result.append({
                        "PGN_DEC": pgn,
                        "PGN_HEX": f"0x{pgn:04X}",
                        "SPNs": [],
                        "message": "No SPN Data Available"
                    })
                    continue

                pgns_with_data += 1
                spn_list = []
                pgn_hex = None

                for row in spn_rows:
                    unique_spns.add(row.SPN_Number)
                    pgn_hex = row.PGN_HEX or f"0x{pgn:04X}"

                    # Calculate bit range
                    start_byte = row.Start_Byte
                    data_length = row.Data_Length_Bytes
                    
                    if data_length == 1:
                        bit_range = str(start_byte)
                    elif data_length == 2:
                        bit_range = f"{start_byte}-{start_byte + 1}"
                    else:
                        bit_range = f"{start_byte}-{start_byte + data_length - 1}"

                    spn_list.append({
                        "SPN": row.SPN_Number,
                        "Description": row.SPN_Description,
                        "Bit_Range": bit_range,
                        "Length_Bits": row.Bit_Length,
                        "Unit": row.Unit
                    })

                result.append({
                    "PGN_DEC": pgn,
                    "PGN_HEX": pgn_hex,
                    "SPNs": spn_list
                })

            return Response({
                "status": "success",
                "filename": file.name,
                "total_unique_spn_count": len(unique_spns),
                "total_pgn_count": len(pgn_list),
                "pgns_with_spn_data": pgns_with_data,
                "pgns_without_spn_data": pgns_without_data,
                "unique_spn_list": sorted(list(unique_spns)),
                "data": result
            })

        except Exception as e:
            return Response({
                'error': 'Failed to analyze file',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


